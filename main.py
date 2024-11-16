from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from pathlib import Path
from decimal import Decimal
from openpyxl.styles import PatternFill
import tkinter as tk
from tkinter import filedialog, messagebox

#Givens
pipeline = "??mm ?? Transmission"
reportedBy = "Global Raymac - ?? - 2023"
masterSheetName = "CC Survey 2023"

#Colors
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
orange_fill = PatternFill(start_color= "FFA500", end_color="FFA500", fill_type="solid")


def main():
    #path of the current directory
    current_dir = Path('.')
    #saving the names of all the files with suffix ".xlsx"
    xlsx_files = [file.name for file in current_dir.iterdir() if file.is_file() and file.suffix == '.xlsx']
    #load master excel workbook
    masterwb = load_workbook('main/master.xlsx')
    print('loaded master workbook')
    #load the master sheet
    masterws = masterwb[masterSheetName]
    print('loaded master worksheet')


    #main loop with the number of files
    for i,file_name in enumerate(xlsx_files):
        #load current excel workbook
        wb = load_workbook(file_name, data_only=True)
        print('loaded secondary workbook')
        #load the main sheet
        ws = wb.active
        #loop for each row
        for row in ws.iter_rows(min_row=6,values_only=True):
            try:
                easting = row[3] if row[3] is not None else None
                northing = row[2] if row[2] is not None else None
                if easting and northing:
                    lat, lon = convert_to_latlon(easting,northing)
                    print('calculated latitude and longitude')
                ground = row[4] if row[4] is not None else None
                print(type(ground))
                top = row[5] if row[5] is not None else None
                print(type(top))
                if ground and top:
                    print("hey")
                    top = float(top)
                    coverage = (ground - top)
                    print('calculated coverage')
                chainage = row[0] if row[0] is not None else None
                comments = row[9] if len(row) > 9 and row[9] is not None else None
                water = row[7] if len(row) > 7 and row[7] is not None else None
                masterws.append([pipeline,chainage,lat,lon,ground,top,water,coverage,comments,reportedBy])
                print('appended data')
                if masterws[masterws.max_row][7].value <= 0.6:
                    masterws[masterws.max_row][7].fill = red_fill
                elif masterws[masterws.max_row][7].value > 0.6 and masterws[masterws.max_row][7].value <= 1.19:
                    masterws[masterws.max_row][7].fill = yellow_fill
            except Exception as e:
                print("An error occurred in row ",e)
        masterws[masterws.max_row][1].fill= orange_fill
        masterwb.save('main/master.xlsx')
        
if __name__ == '__main__':
    main()