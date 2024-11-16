from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from pathlib import Path
from openpyxl.styles import PatternFill
import tkinter as tk
from tkinter import filedialog, messagebox

class ExcelAutomationApp:
    def __init__(self):
        self.selected_files = []  # To store selected file paths
        self.master_file = None  # To store the master file path
        self.root = None
        self.pipeline = None
        self.reported_by = None
        self.sheet_name = None
        self.create_main_window()

    def select_files(self):
        # Open a file dialog to select multiple files
        file_paths = filedialog.askopenfilenames(
            title="Select Excel Files to Handle",
            filetypes=[("Excel Files", "*.xlsx *.xls")],
        )
        if file_paths:
            self.selected_files = list(file_paths)  # Save the paths
            self.update_selected_files_label()
            self.check_all_files_selected()
            messagebox.showinfo("Files Selected", f"Selected files:\n{', '.join(self.selected_files)}")
            self.handleDuplication()
            print("Selected Files:", self.selected_files)
        else:
            messagebox.showwarning("No Files Selected", "Please select at least one file.")

    def select_master_file(self):
        # Open a file dialog to select a single master file
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            self.master_file = file_path  # Save the master file path
            self.update_master_file_label()
            self.check_all_files_selected()
            messagebox.showinfo("Success", f"Opened {file_path} as the Master File")
            self.handleDuplication()
            print("Master File:", self.master_file)
        else:
            messagebox.showwarning("No File Selected", "Please select the master file.")
    
    def update_selected_files_label(self):
        # Update the label showing selected files
        if self.selected_files:
            files_display = ""
            for file in self.selected_files:
                files_display += file[file.rindex('/')+1:] + "\n"
            self.selected_files_label.config(text=f"Selected Files:\n{files_display}")
        else:
            self.selected_files_label.config(text="Selected Files: None")

    def update_master_file_label(self):
        # Update the label showing the master file
        if self.master_file:
            self.master_file_label.config(text=f"Master File:\n{self.master_file[self.master_file.rindex('/')+1:]}")
        else:
            self.master_file_label.config(text="Master File: None")
            
    def handleDuplication(self):
        if self.master_file in self.selected_files:
            self.master_file = None
            self.selected_files.clear()
            self.update_master_file_label()
            self.update_selected_files_label()
            self.check_all_files_selected()
            messagebox.showwarning("Files Error", "One of the Files was Selected as Master and Scraped, Please Reselect the Files.")
    def check_all_files_selected(self):
        if self.selected_files and self.master_file and self.reported_by.get() != "" and self.pipeline.get() != "" and self.sheet_name.get() != "":
            self.process_button.config(state=tk.NORMAL)
        else:
            self.process_button.config(state=tk.DISABLED)
    def process_files(self):
        return

    def create_main_window(self):
        # Create the main GUI window
        self.root = tk.Tk()
        self.root.title("Nada's Automation Script")
        self.root.geometry("400x800")
        
        self.pipeline = tk.StringVar(self.root)
        self.reported_by = tk.StringVar(self.root)
        self.sheet_name = tk.StringVar(self.root)
        # Buttons
        select_files_button = tk.Button(self.root, text="Select Files", command=self.select_files)
        select_files_button.pack(pady=10)

        select_master_file_button = tk.Button(self.root, text="Select Master File", command=self.select_master_file)
        select_master_file_button.pack(pady=10)
        
        self.selected_files_label = tk.Label(self.root, text="Selected Files: None", justify="left", anchor="w")
        self.selected_files_label.pack(pady=10)

        self.master_file_label = tk.Label(self.root, text="Master File: None", justify="left", anchor="w")
        self.master_file_label.pack(pady=10)
        
        input_frame = tk.Frame(self.root)
        input_frame.pack(side="left", padx=20, pady=10, anchor="n")
        
        tk.Label(input_frame, text="Pipeline:").pack(pady=5, anchor="w")
        pipeline_entry = tk.Entry(input_frame, textvariable=self.pipeline)
        pipeline_entry.pack(pady=5, anchor="w")

        tk.Label(input_frame, text="Reported By:").pack(pady=5, anchor="w")
        reported_by_entry = tk.Entry(input_frame, textvariable=self.reported_by)
        reported_by_entry.pack(pady=5, anchor="w")

        tk.Label(input_frame, text="Sheet Name:").pack(pady=5, anchor="w")
        sheet_name_entry = tk.Entry(input_frame, textvariable=self.sheet_name)
        sheet_name_entry.pack(pady=5, anchor="w")

        # Bind input fields to update the process button state
        self.pipeline.trace_add("write", lambda *args: self.check_all_files_selected())
        self.reported_by.trace_add("write", lambda *args: self.check_all_files_selected())
        self.sheet_name.trace_add("write", lambda *args: self.check_all_files_selected())

        
        self.process_button = tk.Button(self.root, text="Process Files", command=self.process_files, state=tk.DISABLED)
        self.process_button.pack(pady=20)

    def run(self):
        # Run the application
        self.root.mainloop()


if __name__ == "__main__":
    # Create an instance of the app and run it
    app = ExcelAutomationApp()
    app.run()
